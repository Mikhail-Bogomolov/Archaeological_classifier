"""Словари значений признаков из Excel Канск 2023."""

from __future__ import annotations

import re
from collections import Counter
from pathlib import Path

import openpyxl

from app.ml.config import FEATURE_SCHEMA, KANSK_TABLES_DIR, OBJECT_CLASSES
from app.ml.table_normalization import (
    NOT_SPECIFIED_VALUES,
    normalize_cell_value,
)

# Старые имена колонок в Excel → актуальные из FEATURE_SCHEMA.
FEATURE_COLUMN_ALIASES: dict[str, list[str]] = {
    "тип_насада": ["тип_насадки"],
    "форма_пера": ["форма_лезвия"],
}


def attribute_key(class_name: str, feature_name: str) -> str:
    return f"{class_name}:{feature_name}"


def head_key(attr_key: str) -> str:
    return attr_key.replace(":", "__")


def from_head_key(head: str) -> str:
    return head.replace("__", ":")


def _parse_attr_key(attr_key: str) -> tuple[str, str]:
    class_name, feature_name = attr_key.split(":", 1)
    return class_name, feature_name


def normalize_feature_value(
    raw: object,
    *,
    class_name: str | None = None,
    feature_name: str | None = None,
) -> str | None:
    if raw is None:
        return None
    text = str(raw).strip().lower()
    text = re.sub(r"\s+", " ", text)
    if text in NOT_SPECIFIED_VALUES:
        return None
    if class_name and feature_name:
        return normalize_cell_value(class_name, feature_name, text)
    return text


def resolve_column_index(headers: list[str], feature_name: str) -> int | None:
    """Индекс колонки признака с учётом старых имён."""
    names = [feature_name, *FEATURE_COLUMN_ALIASES.get(feature_name, [])]
    for name in names:
        if name in headers:
            return headers.index(name)
    return None


def scan_tables(tables_dir: str | Path = KANSK_TABLES_DIR) -> dict[str, Counter]:
    tables_dir = Path(tables_dir)
    counts: dict[str, Counter] = {}

    for class_name in OBJECT_CLASSES:
        xlsx = tables_dir / f"{class_name}.xlsx"
        if not xlsx.is_file():
            continue
        wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            continue
        headers = [str(h).strip() if h is not None else "" for h in header_row]

        for feature_name in FEATURE_SCHEMA.get(class_name, []):
            idx = resolve_column_index(headers, feature_name)
            if idx is None:
                continue
            key = attribute_key(class_name, feature_name)
            counts.setdefault(key, Counter())
            for row in ws.iter_rows(min_row=2, values_only=True):
                if idx >= len(row):
                    continue
                norm = normalize_feature_value(
                    row[idx], class_name=class_name, feature_name=feature_name
                )
                if norm:
                    counts[key][norm] += 1
    return counts


def build_vocab(
    tables_dir: str | Path = KANSK_TABLES_DIR,
    min_count: int = 1,
    min_classes: int = 2,
) -> dict[str, list[str]]:
    """attr_key → отсортированный список классов (минимум min_classes значения)."""
    counts = scan_tables(tables_dir)
    vocab: dict[str, list[str]] = {}

    for key, counter in counts.items():
        counter = _merge_near_duplicate_labels(counter)
        labels = [label for label, c in counter.items() if c >= min_count]
        labels.sort(key=lambda x: (-counter[x], x))
        if len(labels) >= min_classes:
            vocab[key] = labels
    return vocab


def _labels_should_not_merge(a: str, b: str) -> bool:
    """Не сливать разные классы маппинга (ложные дубликаты difflib)."""
    if a == b:
        return False
    if a.startswith("не") and a[2:] == b:
        return True
    if b.startswith("не") and b[2:] == a:
        return True
    distinct = {
        frozenset({"целый", "сломан"}),
        frozenset({"прямой", "изогнутый"}),
        frozenset({"плоское", "объемное"}),
        frozenset({"внутреннее", "внешнее"}),
        frozenset({"да", "нет"}),
        frozenset({"втульчатый", "черешковый"}),
        frozenset({"овальный", "прямоугольный"}),
        frozenset({"кольчато-овальное", "подтреугольное"}),
        frozenset({"бронза", "железо"}),
        frozenset({"заостренный", "расщепленный"}),
        frozenset({"заостренный", "прямой"}),
        frozenset({"прямой", "расщепленный"}),
        frozenset({"округлая", "прямоугольная"}),
        frozenset({"округлая", "фигурная"}),
        frozenset({"прямоугольная", "фигурная"}),
        frozenset({"выделенная", "невыделенная"}),
    }
    return frozenset({a, b}) in distinct


def _merge_near_duplicate_labels(
    counter: Counter,
    *,
    cutoff: float = 0.90,
) -> Counter:
    """Сливает опечатки (difflib) в более частый канон внутри одной головы."""
    from difflib import SequenceMatcher

    labels = sorted(counter.keys(), key=lambda x: (-counter[x], len(x), x))
    canonical: list[str] = []
    mapping: dict[str, str] = {}
    for label in labels:
        matched = None
        for canon in canonical:
            if _labels_should_not_merge(label, canon):
                continue
            if SequenceMatcher(None, label, canon).ratio() >= cutoff:
                matched = canon
                break
        if matched is None:
            canonical.append(label)
            mapping[label] = label
        else:
            mapping[label] = matched
    merged: Counter = Counter()
    for label, count in counter.items():
        merged[mapping[label]] += count
    return merged


def value_to_index(vocab: dict[str, list[str]], attr_key: str, raw: object) -> int:
    from difflib import get_close_matches

    class_name, feature_name = _parse_attr_key(attr_key)
    norm = normalize_feature_value(
        raw, class_name=class_name, feature_name=feature_name
    )
    if norm is None:
        return -1
    labels = vocab.get(attr_key)
    if not labels:
        return -1
    if norm in labels:
        return labels.index(norm)
    candidates = [lab for lab in labels if lab != "другое"]
    close = get_close_matches(norm, candidates, n=1, cutoff=0.90)
    if close:
        return labels.index(close[0])
    if "другое" in labels:
        return labels.index("другое")
    return -1


def ensure_other_bucket(vocab: dict[str, list[str]], tables_dir: str | Path) -> dict[str, list[str]]:
    """Редкие значения сводим в «другое», если встречаются в таблице."""
    counts = scan_tables(tables_dir)
    out: dict[str, list[str]] = {}

    for key, labels in vocab.items():
        class_name, feature_name = _parse_attr_key(key)
        counter = counts.get(key, Counter())
        frequent = set(labels)
        merged = list(labels)
        rare_seen = any(
            normalize_feature_value(v, class_name=class_name, feature_name=feature_name)
            not in frequent
            and normalize_feature_value(v, class_name=class_name, feature_name=feature_name)
            for v, _ in counter.items()
        )
        if rare_seen and "другое" not in merged:
            merged.append("другое")
        out[key] = merged
    return out
