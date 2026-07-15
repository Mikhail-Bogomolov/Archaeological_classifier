"""Датасет фото + признаков из Excel Канск 2023."""

from __future__ import annotations

import random
from pathlib import Path
from typing import Optional

import numpy as np
import openpyxl
import torch
from torch.utils.data import Dataset
from torchvision import transforms

from app.ml.config import (
    FEATURE_SCHEMA,
    KANSK_PHOTOS_DIR,
    KANSK_TABLES_DIR,
    OBJECT_CLASSES,
    USE_TEXTURE_FEATURES,
)
from app.ml.dataset_stats import RowLoadStats
from app.ml.feature_vocab import (
    attribute_key,
    build_vocab,
    ensure_other_bucket,
    resolve_column_index,
    value_to_index,
)
from app.ml.preprocess import item_key_from_filename, load_classifier_rgb
from app.ml.splits import (
    DEFAULT_SPLIT_SEED,
    DEFAULT_TEST_RATIO,
    DEFAULT_VAL_RATIO,
    rows_for_split,
)
from app.ml.texture_features import extract_texture_vector


class KanskFeatureDataset(Dataset):
    CLASS_TO_IDX = {c: i for i, c in enumerate(OBJECT_CLASSES)}

    def __init__(
        self,
        vocab: dict[str, list[str]],
        photos_dir: str | Path = KANSK_PHOTOS_DIR,
        tables_dir: str | Path = KANSK_TABLES_DIR,
        transform: Optional[transforms.Compose] = None,
        split: str = "train",
        val_ratio: float = DEFAULT_VAL_RATIO,
        test_ratio: float = DEFAULT_TEST_RATIO,
        seed: int = DEFAULT_SPLIT_SEED,
        use_texture: bool = USE_TEXTURE_FEATURES,
        cache_texture: bool = True,
    ):
        if split not in ("train", "val", "test"):
            raise ValueError(f"split must be train|val|test, got {split!r}")
        self.photos_dir = Path(photos_dir)
        self.transform = transform
        self.vocab = vocab
        self.attr_keys = sorted(vocab.keys())
        self.use_texture = use_texture
        self.cache_texture = cache_texture
        self._tex_cache: dict[str, np.ndarray] = {}

        rows, load_stats = self._load_rows(Path(tables_dir))
        rows = self._filter_existing(rows)
        if not rows:
            raise FileNotFoundError(
                f"Нет размеченных фото. Проверьте {tables_dir}/*.xlsx"
            )

        self.samples = rows_for_split(
            rows,
            split,
            val_ratio=val_ratio,
            test_ratio=test_ratio,
            seed=seed,
        )
        self.load_stats = load_stats

        labeled = sum(
            1
            for r in self.samples
            for k in self.attr_keys
            if r["targets"].get(k, -1) >= 0
        )
        tex_note = ", texture" if use_texture else ""
        print(
            f"[KanskFeatureDataset] {split}: {len(self.samples)} фото, "
            f"{len(self.attr_keys)} признаков, {labeled} меток{tex_note}"
        )
        for line in load_stats.summary_lines("[KanskFeatureDataset] "):
            if "пропущ" in line.lower() or "  - " in line:
                print(line)

    @classmethod
    def build_vocab(cls, tables_dir: str | Path = KANSK_TABLES_DIR) -> dict[str, list[str]]:
        vocab = build_vocab(tables_dir, min_count=1, min_classes=2)
        return ensure_other_bucket(vocab, tables_dir)

    def _load_rows(self, tables_dir: Path) -> tuple[list[dict], RowLoadStats]:
        result: list[dict] = []
        seen: set[str] = set()
        stats = RowLoadStats()

        for class_name in OBJECT_CLASSES:
            xlsx = tables_dir / f"{class_name}.xlsx"
            if not xlsx.is_file():
                stats.skip(f"нет файла {class_name}.xlsx")
                continue

            wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
            ws = wb.active
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row:
                stats.skip(f"пустой заголовок ({class_name})")
                continue
            headers = [str(h).strip() if h is not None else "" for h in header_row]
            col_index = {name: i for i, name in enumerate(headers)}

            img_col = col_index.get("image_path")
            if img_col is None:
                stats.skip(f"нет image_path ({class_name})")
                continue

            feature_names = FEATURE_SCHEMA.get(class_name, [])

            for row in ws.iter_rows(min_row=2, values_only=True):
                stats.rows_total += 1
                img_rel = str(row[img_col]).strip() if row[img_col] else ""
                if not img_rel:
                    stats.skip("пустой image_path")
                    continue
                img_path = self.photos_dir / Path(img_rel).name
                key = str(img_path).lower()
                if key in seen:
                    stats.skip("дубликат пути")
                    continue
                seen.add(key)

                targets: dict[str, int] = {}
                for feat_name in feature_names:
                    akey = attribute_key(class_name, feat_name)
                    if akey not in self.vocab:
                        stats.skip("признак не в vocab")
                        continue
                    idx = resolve_column_index(headers, feat_name)
                    raw = row[idx] if idx is not None and idx < len(row) else None
                    targets[akey] = value_to_index(self.vocab, akey, raw)

                if not any(v >= 0 for v in targets.values()):
                    stats.skip("нет размеченных признаков")
                    continue

                stats.rows_kept += 1
                result.append({
                    "path": img_path,
                    "class_idx": self.CLASS_TO_IDX[class_name],
                    "class_name": class_name,
                    "item_key": item_key_from_filename(img_path),
                    "targets": targets,
                })
        return result, stats

    def _filter_existing(self, rows: list[dict]) -> list[dict]:
        ok = [r for r in rows if r["path"].is_file()]
        missing = len(rows) - len(ok)
        if missing:
            print(f"[KanskFeatureDataset] Пропущено {missing} отсутствующих фото.")
        return ok

    def _texture_for_path(self, path: Path, pil) -> np.ndarray | None:
        if not self.use_texture:
            return None
        key = str(path)
        if self.cache_texture and key in self._tex_cache:
            return self._tex_cache[key]
        tex = extract_texture_vector(pil)
        if self.cache_texture:
            self._tex_cache[key] = tex
        return tex

    def __len__(self) -> int:
        return len(self.samples)

    def __getitem__(self, idx: int):
        item = self.samples[idx]
        pil = load_classifier_rgb(path=item["path"])
        tex_vec = self._texture_for_path(item["path"], pil)
        try:
            x = self.transform(pil) if self.transform else transforms.ToTensor()(pil)
        finally:
            pil.close()

        class_idx = item["class_idx"]
        one_hot = torch.zeros(len(OBJECT_CLASSES), dtype=torch.float32)
        one_hot[class_idx] = 1.0

        targets = torch.full((len(self.attr_keys),), -1, dtype=torch.long)
        for i, akey in enumerate(self.attr_keys):
            if akey in item["targets"]:
                targets[i] = item["targets"][akey]

        if self.use_texture and tex_vec is not None:
            tex = torch.tensor(tex_vec, dtype=torch.float32)
            return x, one_hot, torch.tensor(class_idx, dtype=torch.long), targets, tex
        return x, one_hot, torch.tensor(class_idx, dtype=torch.long), targets


def collate_features(batch):
    if len(batch[0]) == 5:
        xs, one_hots, class_idxs, targets, texs = zip(*batch)
        return (
            torch.stack(xs),
            torch.stack(one_hots),
            torch.stack(class_idxs),
            torch.stack(targets),
            torch.stack(texs),
        )
    xs, one_hots, class_idxs, targets = zip(*batch)
    return (
        torch.stack(xs),
        torch.stack(one_hots),
        torch.stack(class_idxs),
        torch.stack(targets),
        None,
    )
