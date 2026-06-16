"""
Обработка датасета «Канск 2023» → data/perdataset/kansk_2023/

Источники (только чтение):
  General_data/Канск 2023/Приложение 1_Опись находок из Канска дело 085_1.docx
  General_data/Канск 2023/ФОТО_2_Следственный отдел по Канскому Району УД85/

Имена фото: УД85_{№ упаковки}-{№ предмета}_<ракурс>.JPG

В Excel только признаки, которые извлекаются из полей описи (без полного текста описания).
Пустое значение признака = «не указано» (в тексте описи нет явного упоминания).
"""

from __future__ import annotations

import re
import shutil
from collections import Counter, defaultdict
from pathlib import Path

from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Font

PROJECT_ROOT = Path(__file__).resolve().parents[1]
GENERAL_DATA = PROJECT_ROOT / "General_data"
OUTPUT_ROOT = PROJECT_ROOT / "data" / "perdataset" / "kansk_2023"
PHOTOS_OUT = OUTPUT_ROOT / "photos"

NOT_SPECIFIED = "не указано"

TARGET_CLASSES = [
    "кельты",
    "ножи",
    "удила",
    "наконечники стрел",
    "накладки",
]

CLASS_PATTERNS: list[tuple[str, list[str]]] = [
    ("кельты", [r"кельт"]),
    ("ножи", [r"\bнож\b", r"ножи", r"ножа", r"ножей"]),
    ("удила", [r"удил"]),
    ("наконечники стрел", [r"наконечник"]),
    ("накладки", [r"накладк"]),
]

# Признаки для сети 2 — только из полей описи
FEATURE_COLUMNS: dict[str, list[str]] = {
    "кельты": [
        "материал",
        "сохранность",
        "форма",
        "тулья",
        "ушки",
        "орнамент",
    ],
    "ножи": [
        "материал",
        "сохранность",
        "тип",
        "рукоять",
        "орнамент",
    ],
    "удила": [
        "материал",
        "сохранность",
        "тип_окончания",
        "шарнирность",
    ],
    "наконечники стрел": [
        "материал",
        "сохранность",
        "тип_насадки",
        "форма_лезвия",
        "сечение",
        "ребра",
    ],
    "накладки": [
        "материал",
        "сохранность",
        "форма",
        "крепление",
        "орнамент",
    ],
}


def find_kansk_dir() -> Path:
    for p in GENERAL_DATA.iterdir():
        if p.is_dir() and "канск" in p.name.lower():
            return p
    raise FileNotFoundError("Не найдена папка General_data/Канск 2023")


def find_inventory_docx(kansk_dir: Path) -> Path:
    for p in kansk_dir.iterdir():
        if p.suffix.lower() == ".docx" and "085_1" in p.name:
            return p
    raise FileNotFoundError("Не найден docx Приложение 1_Опись ... 085_1")


def find_photos_dir(kansk_dir: Path) -> Path:
    for p in kansk_dir.iterdir():
        if p.is_dir() and p.name.startswith("ФОТО_2"):
            return p
    raise FileNotFoundError("Не найдена папка ФОТО_2_...")


def detect_class(description: str) -> str | None:
    text = description.lower()
    for class_name, patterns in CLASS_PATTERNS:
        for pat in patterns:
            if re.search(pat, text):
                return class_name
    return None


def photo_prefix(pack_no: str, item_no: str) -> str:
    return f"уд85_{pack_no.strip()}-{item_no.strip()}".lower()


def index_photos(photos_dir: Path) -> dict[str, list[Path]]:
    by_key: dict[str, list[Path]] = defaultdict(list)
    pat = re.compile(r"^уд85_(\d+)-(\d+)", re.IGNORECASE)
    for p in photos_dir.iterdir():
        if not p.is_file():
            continue
        m = pat.match(p.name)
        if m:
            key = f"уд85_{m.group(1)}-{m.group(2)}".lower()
            by_key[key].append(p)
    for key in by_key:
        by_key[key].sort(key=lambda x: x.name.lower())
    return by_key


def _norm(value: str) -> str:
    v = value.strip()
    return v if v else NOT_SPECIFIED


def _first_match(desc: str, mapping: list[tuple[str, str]]) -> str:
    """Вернуть значение по первому совпадению шаблона в описании."""
    d = desc.lower()
    for pattern, label in mapping:
        if pattern in d:
            return label
    return NOT_SPECIFIED


def _ornament(desc: str) -> str:
    d = desc.lower()
    if "без орнамент" in d:
        return "нет"
    if "орнамент" in d or "гравирован" in d or "гравировк" in d:
        return "да"
    return NOT_SPECIFIED


def _ushki(desc: str) -> str:
    d = desc.lower()
    if "без ушек" in d or "без уш" in d:
        return "нет"
    if "двуушков" in d or "ушк" in d:
        return "да"
    return NOT_SPECIFIED


def extract_features(
    class_name: str,
    description: str,
    material: str,
    preservation: str,
) -> dict[str, str]:
    """Извлечение признаков только из явных полей описи."""
    desc = description
    feats: dict[str, str] = {
        "материал": _norm(material),
        "сохранность": _norm(preservation),
    }

    if class_name == "кельты":
        feats["форма"] = _first_match(
            desc,
            [
                ("клиновидн", "клиновидный"),
                ("топоровидн", "топоровидный"),
                ("округл", "округлое сечение"),
                ("квадратн", "квадратная форма"),
            ],
        )
        feats["тулья"] = _first_match(
            desc,
            [
                ("втульчат", "втульчатый"),
                ("втулочн", "втулочный"),
                ("двуушковый с муфтой", "двуушковый с муфтой"),
                ("двуушков", "двуушковый"),
            ],
        )
        feats["ушки"] = _ushki(desc)
        feats["орнамент"] = _ornament(desc)

    elif class_name == "ножи":
        feats["тип"] = _first_match(
            desc,
            [
                ("выпуклообушков", "выпуклообушковый"),
                ("коленчат", "коленчатый"),
                ("прямой", "прямой"),
                ("тесло", "тесло"),
            ],
        )
        feats["рукоять"] = _first_match(
            desc,
            [
                ("без выделенной рукояти", "без выделенной рукояти"),
                ("грибовидн", "грибовидное окончание"),
                ("кольцев", "кольцевое навершие"),
                ("петельчат", "петельчатая рукоять"),
                ("рукоят", "есть"),
            ],
        )
        feats["орнамент"] = _ornament(desc)

    elif class_name == "удила":
        feats["тип_окончания"] = _first_match(
            desc,
            [
                ("шарнирн", "шарнирные"),
                ("кольцевидно-овальн", "кольцевидно-овальное"),
                ("овально-кольчат", "овально-кольчатое"),
                ("трапециевидн", "трапециевидное"),
                ("овальн", "овальное"),
            ],
        )
        feats["шарнирность"] = _first_match(
            desc,
            [("шарнирн", "шарнирные")],
        )

    elif class_name == "наконечники стрел":
        feats["тип_насадки"] = _first_match(
            desc,
            [
                ("втульчат", "втульчатый"),
                ("черешков", "черешковый"),
                ("расщепленн", "расщепленный насад"),
                ("втулк", "втулочный"),
            ],
        )
        feats["форма_лезвия"] = _first_match(
            desc,
            [
                ("пулевидн", "пулевидный"),
                ("шипаст", "шипастый"),
                ("плоск", "плоский"),
                ("треугольн", "треугольный"),
                ("трапециевидн", "трапециевидное лезвие"),
            ],
        )
        feats["сечение"] = _first_match(
            desc,
            [("ромбовидн", "ромбовидное")],
        )
        feats["ребра"] = (
            "да"
            if "ребр" in desc.lower()
            else NOT_SPECIFIED
        )

    elif class_name == "накладки":
        feats["форма"] = _first_match(
            desc,
            [
                ("дисковидн", "дисковидная"),
                ("каплевидн", "каплевидная"),
                ("подовальн", "подовальная"),
                ("ажурн", "ажурная"),
                ("кругл", "круглая"),
            ],
        )
        feats["крепление"] = _first_match(
            desc,
            [
                ("шпеньк", "шпеньки"),
                ("петельк", "петелька"),
                ("клепк", "клепка"),
                ("пронизк", "пронизка"),
            ],
        )
        feats["орнамент"] = _ornament(desc)

    # Заполнить все колонки класса значением «не указано», если не извлекли
    for col in FEATURE_COLUMNS[class_name]:
        feats.setdefault(col, NOT_SPECIFIED)

    return {k: feats[k] for k in FEATURE_COLUMNS[class_name]}


def parse_inventory(docx_path: Path, photos_index: dict[str, list[Path]]) -> list[dict]:
    doc = Document(docx_path)
    table = doc.tables[0]
    records: list[dict] = []

    for row in table.rows[1:]:
        cells = [c.text.strip() for c in row.cells]
        if len(cells) < 10:
            continue

        pack_no = cells[0]
        item_no = cells[2]
        description = cells[4]
        material = cells[7] if len(cells) > 7 else ""
        preservation = cells[9] if len(cells) > 9 else ""

        class_name = detect_class(description)
        if not class_name:
            continue

        key = photo_prefix(pack_no, item_no)
        photo_paths = photos_index.get(key, [])
        if not photo_paths:
            continue

        features = extract_features(class_name, description, material, preservation)

        for photo_path in photo_paths:
            records.append(
                {
                    "photo_key": key,
                    "source_photo": photo_path,
                    "class": class_name,
                    "features": features,
                }
            )

    return records


def safe_filename(name: str) -> str:
    return re.sub(r'[<>:"/\\|?*]', "_", name)


def copy_photos(records: list[dict]) -> list[dict]:
    PHOTOS_OUT.mkdir(parents=True, exist_ok=True)
    updated = []
    used_names: set[str] = set()
    for i, rec in enumerate(records):
        src: Path = rec["source_photo"]
        ext = src.suffix.lower()
        base = safe_filename(src.name.lower())
        out_name = base
        n = 1
        while out_name in used_names:
            out_name = safe_filename(f"{src.stem}_{n}{ext}".lower())
            n += 1
        used_names.add(out_name)
        rel_path = Path("photos") / out_name
        dest = OUTPUT_ROOT / rel_path
        if not dest.exists():
            shutil.copy2(src, dest)
        rec = dict(rec)
        rec["image_path"] = str(rel_path).replace("\\", "/")
        rec["sample_id"] = i + 1
        updated.append(rec)
    return updated


def write_class_workbook(class_name: str, records: list[dict], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = class_name[:31]

    feature_cols = FEATURE_COLUMNS[class_name]
    headers = ["sample_id", "image_path", "класс"] + feature_cols
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for rec in records:
        feats = rec["features"]
        row = [rec["sample_id"], rec["image_path"], rec["class"]]
        row.extend(feats[col] for col in feature_cols)
        ws.append(row)

    wb.save(path)


def write_statistics(records: list[dict], path: Path) -> None:
    wb = Workbook()

    ws0 = wb.active
    ws0.title = "пояснения"
    ws0.append(["поле", "значение"])
    ws0.append(
        [
            "не указано",
            "В тексте описи или в столбце опись нет явного значения для этого признака",
        ]
    )
    ws0.append(
        [
            "image_path",
            "Содержит УД85_{упаковка}-{предмет}_ракурс — связь с опись без отдельных столбцов",
        ]
    )

    ws1 = wb.create_sheet("количество_по_классам")
    ws1.append(["класс", "число_фото", "число_предметов_в_описи", "доля_%"])
    class_photo = Counter(r["class"] for r in records)
    class_items = Counter(f"{r['class']}|{r['photo_key']}" for r in records)
    item_counts = Counter()
    for key in class_items:
        cls = key.split("|")[0]
        item_counts[cls] += 1

    total = sum(class_photo.values())
    for cls in TARGET_CLASSES:
        n = class_photo.get(cls, 0)
        pct = round(100 * n / total, 1) if total else 0
        ws1.append([cls, n, item_counts.get(cls, 0), pct])
    ws1.append(["ИТОГО", total, "", 100.0 if total else 0])

    ws2 = wb.create_sheet("пропорции_признаков")
    ws2.append(["класс", "признак", "значение", "количество", "доля_%"])

    by_class: dict[str, list[dict]] = defaultdict(list)
    for r in records:
        by_class[r["class"]].append(r)

    for cls in TARGET_CLASSES:
        cls_records = by_class.get(cls, [])
        if not cls_records:
            continue
        n = len(cls_records)
        for feat_name in FEATURE_COLUMNS[cls]:
            value_counts: Counter[str] = Counter()
            for rec in cls_records:
                val = rec["features"].get(feat_name, NOT_SPECIFIED)
                value_counts[val] += 1
            for val, cnt in value_counts.most_common():
                ws2.append([cls, feat_name, val, cnt, round(100 * cnt / n, 1)])

    wb.save(path)


def write_master_table(records: list[dict], path: Path) -> None:
    """Сводный файл: лист «все» + отдельный лист на каждый класс."""
    wb = Workbook()
    wb.remove(wb.active)

    ws_all = wb.create_sheet(title="все")
    ws_all.append(["sample_id", "image_path", "класс"])
    for cell in ws_all[1]:
        cell.font = Font(bold=True)
    for rec in records:
        ws_all.append([rec["sample_id"], rec["image_path"], rec["class"]])

    by_class: dict[str, list[dict]] = defaultdict(list)
    for r in records:
        by_class[r["class"]].append(r)

    for cls in TARGET_CLASSES:
        cls_records = by_class.get(cls, [])
        ws = wb.create_sheet(title=cls[:31])
        feature_cols = FEATURE_COLUMNS[cls]
        headers = ["sample_id", "image_path", "класс"] + feature_cols
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for rec in cls_records:
            feats = rec["features"]
            row = [rec["sample_id"], rec["image_path"], rec["class"]]
            row.extend(feats[col] for col in feature_cols)
            ws.append(row)

    wb.save(path)


def main() -> None:
    kansk_dir = find_kansk_dir()
    docx_path = find_inventory_docx(kansk_dir)
    photos_dir = find_photos_dir(kansk_dir)

    print(f"Опись: {docx_path.name}")
    print(f"Фото:  {photos_dir.name} ({len(list(photos_dir.iterdir()))} файлов)")

    photos_index = index_photos(photos_dir)
    records = parse_inventory(docx_path, photos_index)
    print(f"Сопоставлено фото целевых классов: {len(records)}")

    if OUTPUT_ROOT.exists():
        shutil.rmtree(OUTPUT_ROOT)
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    records = copy_photos(records)

    tables_dir = OUTPUT_ROOT / "tables"
    tables_dir.mkdir(exist_ok=True)

    by_class: dict[str, list[dict]] = defaultdict(list)
    for r in records:
        by_class[r["class"]].append(r)

    for cls in TARGET_CLASSES:
        cls_records = by_class.get(cls, [])
        fname = safe_filename(f"{cls}.xlsx")
        write_class_workbook(cls, cls_records, tables_dir / fname)
        print(f"  {cls}: {len(cls_records)} фото -> tables/{fname}")

    write_master_table(records, tables_dir / "all_classes.xlsx")
    write_statistics(records, tables_dir / "statistics.xlsx")

    print(f"\nГотово: {OUTPUT_ROOT}")


if __name__ == "__main__":
    main()
