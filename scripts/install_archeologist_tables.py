"""
Установка таблиц из five_tables_археологи в data/dataset/tables
и пересборка all_classes.xlsx для сети 1.

Запуск:
    py scripts/install_archeologist_tables.py
    py scripts/install_archeologist_tables.py --dry-run
"""

from __future__ import annotations

import argparse
import shutil
import sys
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from app.ml.config import KANSK_TABLES_DIR, OBJECT_CLASSES
SOURCE_DIR = PROJECT_ROOT / "five_tables_археологи"
TABLES_DIR = PROJECT_ROOT / KANSK_TABLES_DIR


def _read_rows(xlsx: Path) -> list[tuple[str, str, str]]:
    """sample_id, image_path, class_name."""
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return []
    headers = [str(h).strip() if h is not None else "" for h in header_row]
    col = {name: i for i, name in enumerate(headers)}

    sid_col = col.get("sample_id")
    img_col = col.get("image_path")
    if img_col is None:
        return []

    rows: list[tuple[str, str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        img = str(row[img_col]).strip() if row[img_col] else ""
        if not img:
            continue
        sid = str(row[sid_col]).strip() if sid_col is not None and row[sid_col] else ""
        rows.append((sid, img, xlsx.stem))
    return rows


def build_all_classes(out_path: Path) -> int:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "все"
    ws.append(["sample_id", "image_path", "класс"])

    total = 0
    for class_name in OBJECT_CLASSES:
        xlsx = TABLES_DIR / f"{class_name}.xlsx"
        if not xlsx.is_file():
            print(f"  пропуск: нет {xlsx.name}")
            continue
        for sid, img, cls in _read_rows(xlsx):
            ws.append([sid, img, cls])
            total += 1

    wb.save(out_path)
    return total


def main() -> None:
    parser = argparse.ArgumentParser(description="Установка таблиц археологов")
    parser.add_argument(
        "--source",
        type=Path,
        default=SOURCE_DIR,
        help="Папка с xlsx (по умолчанию five_tables_археологи)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Только показать, что будет скопировано",
    )
    args = parser.parse_args()

    if not args.source.is_dir():
        raise SystemExit(f"Не найдена папка: {args.source}")

    TABLES_DIR.mkdir(parents=True, exist_ok=True)

    print(f"Источник: {args.source}")
    print(f"Назначение: {TABLES_DIR}")
    for class_name in OBJECT_CLASSES:
        src = args.source / f"{class_name}.xlsx"
        dst = TABLES_DIR / f"{class_name}.xlsx"
        if not src.is_file():
            raise SystemExit(f"Нет файла: {src}")
        if args.dry_run:
            print(f"  [dry-run] {src.name} -> {dst}")
        else:
            shutil.copy2(src, dst)
            print(f"  скопировано: {src.name}")

    all_classes = TABLES_DIR / "all_classes.xlsx"
    if args.dry_run:
        n = sum(len(_read_rows(args.source / f"{c}.xlsx")) for c in OBJECT_CLASSES)
        print(f"  [dry-run] all_classes.xlsx: {n} строк")
    else:
        n = build_all_classes(all_classes)
        print(f"  all_classes.xlsx: {n} строк")

    print("\nДальше:")
    print("  py -m app.ml.train_feature_classifier --verify-only")
    print("  py -m app.ml.train_feature_classifier --epochs 30 --batch-size 16")
    print("  py -m app.ml.train_classifier --epochs 30 --batch-size 16  # опционально")


if __name__ == "__main__":
    main()
