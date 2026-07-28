"""
Аудит проблемных голов сети 2: удила:сохранность, ножи:тип, ножи:орнамент.

Запуск:
    py scripts/audit_critical_feature_heads.py
    py scripts/audit_critical_feature_heads.py --json-out reports/feature_heads_audit.json
"""

from __future__ import annotations

import argparse
import json
import sys
from collections import Counter, defaultdict
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import openpyxl

from app.ml.config import FEATURE_SCHEMA, KANSK_TABLES_DIR, OBJECT_CLASSES
from app.ml.feature_vocab import resolve_column_index, scan_tables
from app.ml.table_normalization import normalize_cell_value

AUDIT_HEADS = [
    ("удила", "сохранность"),
    ("ножи", "тип"),
    ("ножи", "орнамент"),
]


def _load_udila_rows() -> list[dict[str, str]]:
    xlsx = Path(KANSK_TABLES_DIR) / "удила.xlsx"
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return []
    headers = [str(h).strip() if h is not None else "" for h in header_row]
    col_soh = resolve_column_index(headers, "сохранность")
    col_name = headers.index("название") if "название" in headers else None
    col_num = headers.index("номер") if "номер" in headers else None
    rows: list[dict[str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        raw_soh = row[col_soh] if col_soh is not None and col_soh < len(row) else None
        norm = normalize_cell_value("удила", "сохранность", raw_soh)
        if norm is None:
            continue
        rows.append(
            {
                "номер": str(row[col_num] if col_num is not None else ""),
                "название": str(row[col_name] if col_name is not None else ""),
                "raw": str(raw_soh).strip() if raw_soh is not None else "",
                "normalized": norm,
            }
        )
    return rows


def _knife_tip_raw_mapping() -> dict[str, list[str]]:
    xlsx = Path(KANSK_TABLES_DIR) / "ножи.xlsx"
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    headers = [str(h).strip() if h is not None else "" for h in header_row or []]
    col_tip = resolve_column_index(headers, "тип")
    by_norm: dict[str, list[str]] = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if col_tip is None or col_tip >= len(row):
            continue
        raw = row[col_tip]
        if raw is None or str(raw).strip() == "":
            continue
        raw_s = str(raw).strip()
        norm = normalize_cell_value("ножи", "тип", raw)
        key = norm if norm else "(drop)"
        if raw_s not in by_norm[key]:
            by_norm[key].append(raw_s)
    return dict(by_norm)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--json-out", default="reports/feature_heads_audit.json")
    args = parser.parse_args()

    table_counts = scan_tables()
    payload: dict[str, object] = {
        "train_counts": {
            f"{c}:{f}": dict(table_counts.get(f"{c}:{f}", Counter()))
            for c, f in AUDIT_HEADS
        },
    }

    udila_rows = _load_udila_rows()
    udila_by_norm = Counter(r["normalized"] for r in udila_rows)
    udila_samples = {
        norm: [r for r in udila_rows if r["normalized"] == norm][:8]
        for norm in sorted(udila_by_norm.keys())
    }
    payload["удила:сохранность"] = {
        "train_normalized_counts": dict(udila_by_norm),
        "raw_to_normalized_samples": udila_samples,
        "note": (
            "Проверить визуально: целый набор vs обломан — "
            "если на фото неразличимо, голова непригодна для автоматики."
        ),
    }

    tip_map = _knife_tip_raw_mapping()
    dropped = tip_map.get("(drop)", [])
    payload["ножи:тип"] = {
        "raw_strings_by_normalized_class": tip_map,
        "dropped_raw_count": len(dropped),
        "dropped_examples": dropped[:15],
        "note": (
            "листовидный/прямолезвийный → прямой; "
            "выпуклообушковый/змейчатообушковый/коленчатый → изогнутый."
        ),
    }

    orn_counts = dict(table_counts.get("ножи:орнамент", Counter()))
    payload["ножи:орнамент"] = {
        "train_normalized_counts": orn_counts,
        "note": (
            "Бинарный да/нет; слабые метрики на test — "
            "смотреть heatmaps в reports/feature_test_heatmaps/."
        ),
    }

    out = Path(args.json_out)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    print("=== Аудит критических голов (train tables) ===\n")
    for c, f in AUDIT_HEADS:
        key = f"{c}:{f}"
        counts = table_counts.get(key, Counter())
        print(f"{key}: {dict(counts)}")
    print(f"\nудила:сохранность normalized: {dict(udila_by_norm)}")
    print(f"ножи:тип dropped raw (не в обучении): {len(dropped)} строк")
    print(f"\nОтчёт: {out}")


if __name__ == "__main__":
    main()
