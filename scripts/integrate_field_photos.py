"""
Добавление полевых фото из «Новая папка/data» в data/dataset/.

Для каждого снимка:
  - копия в data/dataset/photos/
  - строка в tables/<класс>.xlsx
  - признаки из описи УД85 (docx) по номеру упаковки-предмета (уд85 18-1)

Запуск:
    python scripts/integrate_field_photos.py
    python scripts/integrate_field_photos.py --class накладки
    python scripts/integrate_field_photos.py --dry-run
"""

from __future__ import annotations

import argparse
import re
import shutil
import sys
from pathlib import Path

from docx import Document
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

PROJECT_ROOT = Path(__file__).resolve().parents[1]
FIELD_DATA_DIR = PROJECT_ROOT / "Новая папка" / "data"
DATASET_ROOT = PROJECT_ROOT / "data" / "dataset"
PHOTOS_OUT = DATASET_ROOT / "photos"
TABLES_DIR = DATASET_ROOT / "tables"

sys.path.insert(0, str(PROJECT_ROOT))

from scripts.process_kansk_dataset import (  # noqa: E402
    FEATURE_COLUMNS,
    NOT_SPECIFIED,
    detect_class,
    extract_features,
    find_inventory_docx,
    find_kansk_dir,
    safe_filename,
)

SUPPORTED_FIELD_CLASSES = ("удила", "накладки")

CLASS_HINTS: dict[str, tuple[str, ...]] = {
    "удила": ("удил",),
    "накладки": ("накладк", "бляшк", "бляха", "пуговиц", "нашивк", "ворворк"),
}

JUNK_PATTERNS = (
    r"тест",
    r"\bуд81\b",
)


def load_inventory() -> dict[str, dict]:
    doc = Document(find_inventory_docx(find_kansk_dir()))
    index: dict[str, dict] = {}
    for row in doc.tables[0].rows[1:]:
        cells = [c.text.strip() for c in row.cells]
        if len(cells) < 10:
            continue
        key = f"{cells[0].strip()}-{cells[2].strip()}"
        index[key] = {
            "description": cells[4],
            "material": cells[7] if len(cells) > 7 else "",
            "preservation": cells[9] if len(cells) > 9 else "",
            "inventory_class": detect_class(cells[4]),
        }
    return index


def is_junk_description(text: str) -> bool:
    lower = text.strip().lower()
    return any(re.search(pat, lower) for pat in JUNK_PATTERNS)


def parse_item_key(description: str) -> str | None:
    m = re.search(r"уд85\s+(\d+)\s*-\s*(\d+)", description, re.IGNORECASE)
    if not m:
        return None
    return f"{m.group(1)}-{m.group(2)}"


def parse_field_class(description: str) -> str | None:
    lower = description.strip().lower()
    if lower.startswith("наконечники"):
        return "наконечники стрел"
    for cls in SUPPORTED_FIELD_CLASSES:
        if lower.startswith(cls):
            return cls
    return None


def load_field_pairs(classes: tuple[str, ...]) -> list[dict]:
    pairs: list[dict] = []
    for txt_path in sorted(FIELD_DATA_DIR.glob("data_*.txt")):
        stamp = txt_path.stem.split("_", 1)[1]
        img_path = FIELD_DATA_DIR / f"img_{stamp}.jpg"
        if not img_path.is_file():
            continue
        description = txt_path.read_text(encoding="utf-8").strip()
        if is_junk_description(description):
            continue
        field_class = parse_field_class(description)
        if field_class not in classes:
            continue
        item_key = parse_item_key(description)
        if item_key is None:
            continue
        pairs.append(
            {
                "stamp": stamp,
                "item_key": item_key,
                "description": description,
                "field_class": field_class,
                "image_path": img_path,
            }
        )
    return pairs


def _description_matches_class(description: str, field_class: str) -> bool:
    lower = description.lower()
    return any(hint in lower for hint in CLASS_HINTS.get(field_class, ()))


def features_for_field_photo(
    item_key: str,
    inventory: dict[str, dict],
    field_class: str,
) -> dict[str, str]:
    feat_cols = FEATURE_COLUMNS[field_class]
    inv = inventory.get(item_key)
    if inv is None:
        return {col: NOT_SPECIFIED for col in feat_cols}

    desc = inv["description"]
    mat = inv["material"]
    pres = inv["preservation"]
    inv_cls = inv["inventory_class"]

    if inv_cls == field_class or _description_matches_class(desc, field_class):
        return extract_features(field_class, desc, mat, pres)

    feats = {col: NOT_SPECIFIED for col in feat_cols}
    feats["материал"] = mat.strip().lower() if mat.strip() else NOT_SPECIFIED
    feats["сохранность"] = pres.strip().lower() if pres.strip() else NOT_SPECIFIED
    return feats


def next_sample_id(xlsx_path: Path) -> int:
    if not xlsx_path.is_file():
        return 1
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    max_id = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] is not None:
            try:
                max_id = max(max_id, int(row[0]))
            except (TypeError, ValueError):
                pass
    wb.close()
    return max_id + 1


def existing_field_stamps(xlsx_path: Path) -> set[str]:
    if not xlsx_path.is_file():
        return set()
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    stamps: set[str] = set()
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
        path = str(row[0] or "")
        m = re.search(r"_field_(\d+)\.jpg$", path, re.IGNORECASE)
        if m:
            stamps.add(m.group(1))
    wb.close()
    return stamps


def ensure_class_workbook(path: Path, class_name: str) -> None:
    if path.is_file():
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    feat_cols = FEATURE_COLUMNS[class_name]
    headers = ["sample_id", "image_path", "класс", *feat_cols]
    wb = Workbook()
    ws = wb.active
    ws.title = class_name[:31]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    wb.save(path)


def integrate_class(class_name: str, inventory: dict[str, dict], *, dry_run: bool) -> int:
    xlsx_path = TABLES_DIR / f"{class_name}.xlsx"
    feat_cols = FEATURE_COLUMNS[class_name]
    pairs = [p for p in load_field_pairs((class_name,)) if p["field_class"] == class_name]
    already = existing_field_stamps(xlsx_path)
    to_add = [p for p in pairs if p["stamp"] not in already]

    print(f"\n=== {class_name} ===")
    print(f"Полевых фото: {len(pairs)} | уже в датасете: {len(already)} | добавить: {len(to_add)}")
    if not to_add:
        return 0

    PHOTOS_OUT.mkdir(parents=True, exist_ok=True)
    TABLES_DIR.mkdir(parents=True, exist_ok=True)
    ensure_class_workbook(xlsx_path, class_name)

    sample_id = next_sample_id(xlsx_path)
    new_rows: list[list] = []

    for pair in sorted(to_add, key=lambda p: p["stamp"]):
        item_key = pair["item_key"]
        stamp = pair["stamp"]
        feats = features_for_field_photo(item_key, inventory, class_name)
        out_name = safe_filename(f"уд85_{item_key}_field_{stamp}.jpg")
        rel_path = f"photos/{out_name}"

        inv = inventory.get(item_key)
        inv_note = inv["inventory_class"] if inv else "нет в описи"
        feat_summary = ", ".join(feats[col] for col in feat_cols)
        print(f"  + {out_name} | опись {item_key} ({inv_note}) | {feat_summary}")

        if dry_run:
            sample_id += 1
            continue

        shutil.copy2(pair["image_path"], DATASET_ROOT / rel_path)
        row = [sample_id, rel_path, class_name]
        row.extend(feats[col] for col in feat_cols)
        new_rows.append(row)
        sample_id += 1

    if dry_run or not new_rows:
        return len(to_add) if dry_run else 0

    wb = load_workbook(xlsx_path)
    ws = wb.active
    for row in new_rows:
        ws.append(row)
    wb.save(xlsx_path)
    print(f"Готово: +{len(new_rows)} фото в tables/{class_name}.xlsx")
    return len(new_rows)


def integrate(classes: tuple[str, ...], *, dry_run: bool = False) -> None:
    inventory = load_inventory()
    total = 0
    for class_name in classes:
        total += integrate_class(class_name, inventory, dry_run=dry_run)
    if dry_run:
        print("\ndry-run: файлы не изменены.")
    else:
        print(f"\nИтого добавлено: {total} фото.")


def main() -> None:
    parser = argparse.ArgumentParser(description="Интеграция полевых фото в data/dataset/")
    parser.add_argument(
        "--class",
        dest="classes",
        action="append",
        choices=SUPPORTED_FIELD_CLASSES,
        help="Класс (можно указать несколько раз). По умолчанию — все.",
    )
    parser.add_argument("--dry-run", action="store_true", help="Только показать план")
    args = parser.parse_args()
    classes = tuple(args.classes) if args.classes else SUPPORTED_FIELD_CLASSES
    integrate(classes, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
