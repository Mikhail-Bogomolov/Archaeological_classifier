"""
Применить «Маппинг признаков.md» к data/dataset/tables/*.xlsx
и пересобрать all_classes.xlsx.

Запуск:
    py scripts/apply_feature_mapping.py
    py scripts/apply_feature_mapping.py --dry-run
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from app.ml.config import FEATURE_SCHEMA, KANSK_TABLES_DIR, OBJECT_CLASSES
from app.ml.table_normalization import (
    DROP_FEATURE_COLUMNS,
    coarse_sechenie,
    is_excluded_row_class,
    is_not_specified,
    normalize_cell_value,
)
from scripts.install_archeologist_tables import build_all_classes

TABLES_DIR = PROJECT_ROOT / KANSK_TABLES_DIR
META_COLUMNS = frozenset({"sample_id", "image_path", "класс", "номер", "название"})


def _headers(ws) -> list[str]:
    return [str(c.value).strip() if c.value else "" for c in ws[1]]


def apply_table(path: Path, class_name: str, dry_run: bool) -> dict[str, int]:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = _headers(ws)
    stats = {"rows_drop": 0, "cells": 0, "cols_drop": 0}

    class_col = headers.index("класс") + 1 if "класс" in headers else None
    sech_idx = headers.index("сечение") if "сечение" in headers else None
    rebra_idx = headers.index("ребра") if "ребра" in headers else None

    # Снизу вверх: удаление строк-примесей
    for row in range(ws.max_row, 1, -1):
        if class_col is None:
            break
        row_class = ws.cell(row, class_col).value
        if is_excluded_row_class(class_name, row_class):
            stats["rows_drop"] += 1
            if not dry_run:
                ws.delete_rows(row, 1)

    headers = _headers(ws)
    sech_idx = headers.index("сечение") if "сечение" in headers else None
    rebra_idx = headers.index("ребра") if "ребра" in headers else None

    # Merge сечение+ребра для наконечников
    if class_name == "наконечники стрел" and sech_idx is not None:
        for row in range(2, ws.max_row + 1):
            sec_val = ws.cell(row, sech_idx + 1).value
            reb_val = (
                ws.cell(row, rebra_idx + 1).value if rebra_idx is not None else None
            )
            merged = coarse_sechenie(sec_val, reb_val)
            new_val = merged if merged is not None else "не указано"
            old = sec_val
            if str(old).strip().lower() != str(new_val).strip().lower():
                stats["cells"] += 1
                if not dry_run:
                    ws.cell(row, sech_idx + 1).value = new_val

    # Нормализация ячеек по FEATURE_SCHEMA (+ колонки, которые ещё есть)
    schema = FEATURE_SCHEMA.get(class_name, [])
    drop = DROP_FEATURE_COLUMNS.get(class_name, frozenset())
    headers = _headers(ws)

    for col, header in enumerate(headers, start=1):
        if not header or header in META_COLUMNS:
            continue
        if header in drop:
            continue
        if header not in schema and header != "материал":
            # материал у накладок уже в drop; прочие лишние колонки трогаем только schema
            if header not in schema:
                continue
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row, col)
            old = cell.value
            if old is None or is_not_specified(old):
                if not dry_run:
                    cell.value = "не указано"
                continue
            new = normalize_cell_value(class_name, header, old)
            if new is None:
                stats["cells"] += 1
                if not dry_run:
                    cell.value = "не указано"
                continue
            if str(old).strip() != new:
                stats["cells"] += 1
                if not dry_run:
                    cell.value = new

    # Удаление колонок из DROP (+ пустые хвосты)
    headers = _headers(ws)
    for header in list(headers):
        if header in drop or (not header and header != "0"):
            if header in drop:
                col = headers.index(header) + 1
                stats["cols_drop"] += 1
                if not dry_run:
                    ws.delete_cols(col, 1)
                headers = _headers(ws)

    # Убрать пустые колонки справа
    headers = _headers(ws)
    for col in range(len(headers), 0, -1):
        if not headers[col - 1]:
            if not dry_run:
                ws.delete_cols(col, 1)
            stats["cols_drop"] += 1

    if not dry_run:
        # выставить класс строки = имя таблицы (после фильтрации примесей)
        headers = _headers(ws)
        if "класс" in headers:
            cidx = headers.index("класс") + 1
            for row in range(2, ws.max_row + 1):
                ws.cell(row, cidx).value = class_name
        wb.save(path)
    return stats


def main() -> None:
    parser = argparse.ArgumentParser(description="Применить маппинг признаков к таблицам")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--tables", type=Path, default=TABLES_DIR)
    args = parser.parse_args()

    total = {"rows_drop": 0, "cells": 0, "cols_drop": 0}
    for class_name in OBJECT_CLASSES:
        xlsx = args.tables / f"{class_name}.xlsx"
        if not xlsx.is_file():
            print(f"пропуск: {xlsx}")
            continue
        st = apply_table(xlsx, class_name, args.dry_run)
        for k in total:
            total[k] += st[k]
        print(
            f"{class_name}: ячеек={st['cells']}, строк_удалено={st['rows_drop']}, "
            f"колонок_удалено={st['cols_drop']}"
        )

    if not args.dry_run:
        n = build_all_classes(args.tables / "all_classes.xlsx")
        print(f"\nall_classes.xlsx: {n} строк")

    mode = "[dry-run] " if args.dry_run else ""
    print(
        f"\n{mode}Итого: ячеек={total['cells']}, строк_удалено={total['rows_drop']}, "
        f"колонок_удалено={total['cols_drop']}"
    )


if __name__ == "__main__":
    main()
