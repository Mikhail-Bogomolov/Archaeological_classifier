"""
Нормализация таблиц five_tables_археологи: синонимы, укрупнение форма_пера/крепление,
исправление image_path, единый регистр сохранности/материала.

Запуск:
    py scripts/normalize_archeologist_tables.py
    py scripts/normalize_archeologist_tables.py --dry-run
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from app.ml.config import FEATURE_SCHEMA, OBJECT_CLASSES
from app.ml.table_normalization import (
    fix_image_path,
    is_not_specified,
    normalize_cell_value,
)

SOURCE_DIR = PROJECT_ROOT / "five_tables_археологи"
PHOTOS_DIR = PROJECT_ROOT / "data" / "dataset" / "photos"

SKIP_COLUMNS = frozenset({"sample_id", "image_path", "номер", "название", "класс"})


def normalize_workbook(path: Path, dry_run: bool) -> dict[str, int]:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    stats = {"cells": 0, "paths": 0, "skipped_missing": 0}

    img_col = headers.index("image_path") + 1 if "image_path" in headers else None

    for row in range(2, ws.max_row + 1):
        if img_col:
            raw_path = ws.cell(row, img_col).value
            fixed = fix_image_path(raw_path, PHOTOS_DIR)
            if fixed and str(raw_path).strip() != fixed:
                stats["paths"] += 1
                if not dry_run:
                    ws.cell(row, img_col).value = fixed
            elif fixed and not (PHOTOS_DIR / fixed).is_file():
                stats["skipped_missing"] += 1

        class_name = path.stem
        for col, header in enumerate(headers, start=1):
            if not header or header in SKIP_COLUMNS:
                continue
            if header not in FEATURE_SCHEMA.get(class_name, []) and header not in (
                "материал",
                "сохранность",
            ):
                continue
            cell = ws.cell(row, col)
            old = cell.value
            if old is None or is_not_specified(old):
                if not dry_run:
                    cell.value = "не указано"
                continue
            new = normalize_cell_value(class_name, header, old)
            if new is None:
                if not dry_run:
                    cell.value = "не указано"
                stats["cells"] += 1
                continue
            if _clean_display(old) != new:
                stats["cells"] += 1
                if not dry_run:
                    cell.value = new

    if not dry_run:
        wb.save(path)
    return stats


def _clean_display(raw: object) -> str:
    import re

    return re.sub(r"\s+", " ", str(raw).strip().lower())


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--source", type=Path, default=SOURCE_DIR)
    args = parser.parse_args()

    total = {"cells": 0, "paths": 0, "skipped_missing": 0}
    for class_name in OBJECT_CLASSES:
        xlsx = args.source / f"{class_name}.xlsx"
        if not xlsx.is_file():
            print(f"пропуск: {xlsx}")
            continue
        st = normalize_workbook(xlsx, args.dry_run)
        for k in total:
            total[k] += st[k]
        print(
            f"{class_name}: ячеек={st['cells']}, путей={st['paths']}, "
            f"нет фото={st['skipped_missing']}"
        )

    mode = "[dry-run] " if args.dry_run else ""
    print(
        f"\n{mode}Итого: ячеек={total['cells']}, путей={total['paths']}, "
        f"строк без фото={total['skipped_missing']}"
    )
    if not args.dry_run:
        print("\nДальше: py scripts/install_archeologist_tables.py")


if __name__ == "__main__":
    main()
